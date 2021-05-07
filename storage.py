import pandas as pd
import requests
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill
import smtplib
from email.header import Header
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email import utils
from email import encoders
import os
import datetime

today = datetime.datetime.now()


def post_message(token, channel, text):
    requests.post("https://slack.com/api/chat.postMessage",
                  headers={"Authorization": "Bearer " + token},
                  data={"channel": channel, "text": text}
                  )


channel_name = "#storage_monthly"
slack_token = 'xoxb-'

ToUser = ['abcd@abc.com']
CcUser = ['abcd@abc.com']
FromUser = 'abcd123@gmail.com'
Passwd = 'abcd123'
Server = 'smtp.gmail.com'
Port = 587
Subject = '스토리지 사용량 ' + str(today.month) + '월'
allUser = ToUser + CcUser
def send_email(from_user, to_user, cc_users, subject, textfile, attach):
    COMMASPACE = ', '
    msg = MIMEMultipart('alternative')
    msg['FROM'] = from_user
    msg['To'] = COMMASPACE.join(to_user)
    msg['Cc'] = COMMASPACE.join(cc_users)
    msg['Subject'] = Header(s=subject, charset='utf-8')
    msg['Date'] = utils.formatdate(localtime=1)
    fp = open(textfile, 'rb')
    msg.attach(MIMEText(fp.read().decode('utf8', 'ignore')))
    fp.close()
    if (attach != None):
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(open(attach, 'rb').read())
        encoders.encode_base64(part)
        attachfilename = os.path.basename(attach)
        part.add_header('Content-Disposition', 'attachment', filename=attachfilename)
        msg.attach(part)
        print(attach)
    try:
        smtp = smtplib.SMTP(Server, Port)
        try:
            smtp.starttls()
            smtp.login(FromUser, Passwd)
            smtp.sendmail(from_user, allUser, msg.as_string())
            print("[OK] send mail")
            post_message(slack_token, channel_name, '메일 전송 완료')
        except Exception as e:

            print("[Error] Fail to send mail")
            print(e)
        finally:
            smtp.quit()
    except Exception as e:
        print(e)
        print("[Error] could no connect")
    return False


df_usage = pd.read_csv('usage.txt', header=None,
                       delim_whitespace=True, names=['storage', 'LUN_NAME_tgt', 'SIZE(GB)'])
df_pool = pd.read_csv('pool.txt', header=None, delim_whitespace=True, )

WB = load_workbook('storage_usage.xlsx')
WS = WB.active

yellowFill = PatternFill(start_color='FFF2CC',
                         end_color='FFF2CC',
                         fill_type='solid')
alignment = Alignment(horizontal='center')

for i in range(len(df_usage)):
    WS.cell(row=i + 2, column=1).value = df_usage.loc[i, 'storage']
    WS.cell(row=i + 2, column=2).value = df_usage.loc[i, 'LUN_NAME_tgt']
    WS.cell(row=i + 2, column=3).value = int(df_usage.loc[i, 'SIZE(GB)'])
    i += 1

ds8870_sum_loc = str(len(df_usage) + 2)
ds8700_sum_loc = str(len(df_usage) + 3)
ds8870_merge = 'A' + ds8870_sum_loc + ':B' + ds8870_sum_loc
ds8700_merge = 'A' + ds8700_sum_loc + ':B' + ds8700_sum_loc
last_loc = str(len(df_usage) + 1)
WS['C' + ds8870_sum_loc] = '=SUMIF(A2:A'+last_loc+',"DS8870",C2:C'+last_loc+')/1000&" TB"'
WS['C' + ds8700_sum_loc] = '=SUMIF(A2:A'+last_loc+',"DS8700",C2:C'+last_loc+')/1000&" TB"'
WS['C' + ds8870_sum_loc].fill = yellowFill
WS['C' + ds8700_sum_loc].fill = yellowFill
WS.merge_cells(ds8870_merge)
WS.merge_cells(ds8700_merge)

WS['A' + ds8870_sum_loc] = 'DS8870'
WS['A' + ds8700_sum_loc] = 'DS8700'
WS['A' + ds8870_sum_loc].fill = yellowFill
WS['A' + ds8700_sum_loc].fill = yellowFill

WS.cell(row=int(ds8870_sum_loc), column=1).alignment = alignment
WS.cell(row=int(ds8700_sum_loc), column=1).alignment = alignment
WS['j2'] = '=C' + ds8870_sum_loc
WS['j5'] = '=C' + ds8700_sum_loc
WS['h2'] = df_pool.loc[0, 4]
WS['h3'] = df_pool.loc[1, 4]
WS['h5'] = df_pool.loc[2, 4]
WS['h6'] = df_pool.loc[3, 4]
WS['m1'] = str(today.month) + '월'
localfilename = 'storage_usage_'+str(today.strftime("%Y%m%d"))+'.xlsx'
WB.save(localfilename)
send_email(FromUser, ToUser, CcUser, Subject, 'contents.txt', 'localfilename)
